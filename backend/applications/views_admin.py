from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status as http_status
from rest_framework.pagination import PageNumberPagination

from django.db.models import Q, Avg, Count, FloatField, F, Case, When, Value
from django.db.models.functions import Coalesce
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.http import HttpResponse

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Application, ApplicationScore, ResultNotificationSettings
from .permissions import IsInstructorOrStaff
from .serializers import (
    AdminApplicationSerializer,
    AdminApplicationStatusUpdateSerializer,
    ApplicationScoreSerializer,
    ApplicationScoreUpsertSerializer,
    AdminDecisionFinalizeSerializer,
    ResultNotificationSettingsSerializer,
    PersonalInterviewScheduleSerializer,
)


class AdminPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100


class AdminApplicationListView(APIView):
    permission_classes = [IsInstructorOrStaff]

    def get(self, request):
        qs = (
            Application.objects
            .select_related("user")
            .prefetch_related("scores__reviewer")
            .order_by("-updated_at")
        )

        st = request.query_params.get("status")
        tr = request.query_params.get("track")
        q = request.query_params.get("q")

        if st:
            qs = qs.filter(status=st)
        if tr:
            qs = qs.filter(track=tr)
        if q:
            qs = qs.filter(
                Q(user__email__icontains=q) |
                Q(user__name__icontains=q) |
                Q(user__student_id__icontains=q)
            )

        # ✅ 평균/카운트 annotate (서류/면접 각각)
        qs = qs.annotate(
            doc_avg=Coalesce(
                Avg("scores__score1", filter=Q(scores__kind="DOC"), output_field=FloatField()) +
                Avg("scores__score2", filter=Q(scores__kind="DOC"), output_field=FloatField()) +
                Avg("scores__score3", filter=Q(scores__kind="DOC"), output_field=FloatField()),
                None
            ),
            interview_avg=Coalesce(
                Avg("scores__score1", filter=Q(scores__kind="INTERVIEW"), output_field=FloatField()) +
                Avg("scores__score2", filter=Q(scores__kind="INTERVIEW"), output_field=FloatField()) +
                Avg("scores__score3", filter=Q(scores__kind="INTERVIEW"), output_field=FloatField()),
                None
            ),
            doc_count=Count("scores", filter=Q(scores__kind="DOC"), distinct=True),
            interview_count=Count("scores", filter=Q(scores__kind="INTERVIEW"), distinct=True),
        ).annotate(
            # ✅ 총점 평균: (doc_avg + interview_avg) / 2  (둘 다 있을 때만)
            total_avg=Case(
                When(
                    doc_avg__isnull=False,
                    interview_avg__isnull=False,
                    then=(F("doc_avg") + F("interview_avg")) / Value(2.0),
                ),
                default=None,
                output_field=FloatField(),
            )
        )

        # ✅ 정렬
        sort = request.query_params.get("sort")
        if sort == "TOTAL_DESC":
            qs = qs.order_by(F("total_avg").desc(nulls_last=True), "-updated_at")
        elif sort == "TOTAL_ASC":
            qs = qs.order_by(F("total_avg").asc(nulls_last=True), "-updated_at")
        else:
            qs = qs.order_by("-updated_at")

        paginator = AdminPagination()
        page = paginator.paginate_queryset(qs, request)

        # SerializerMethodField에서 prefetch된 scores를 빠르게 쓰기 위해 붙여줌
        for obj in page:
            obj._prefetched_scores = list(obj.scores.all())

        ser = AdminApplicationSerializer(page, many=True)
        return paginator.get_paginated_response({
            "ok": True,
            "results": ser.data,
        })


class AdminApplicationDetailView(APIView):
    permission_classes = [IsInstructorOrStaff]

    def get(self, request, app_id: int):
        app = get_object_or_404(Application.objects.select_related("user"), id=app_id)
        ser = AdminApplicationSerializer(app)
        return Response({"ok": True, "application": ser.data}, status=200)


class AdminApplicationStatusUpdateView(APIView):
    permission_classes = [IsInstructorOrStaff]

    def patch(self, request, app_id: int):
        app = get_object_or_404(Application.objects.select_related("user"), id=app_id)

        ser = AdminApplicationStatusUpdateSerializer(app, data=request.data, partial=True)
        if not ser.is_valid():
            return Response({"ok": False, "errors": ser.errors}, status=http_status.HTTP_400_BAD_REQUEST)

        ser.save()
        return Response({"ok": True, "status": app.status}, status=200)


class AdminApplicationScoresView(APIView):
    permission_classes = [IsInstructorOrStaff]

    def get(self, request, app_id: int):
        app = get_object_or_404(Application, id=app_id)

        qs = (
            ApplicationScore.objects
            .select_related("reviewer")
            .filter(application=app)
            .order_by("kind", "reviewer_id")
        )

        doc = qs.filter(kind="DOC")
        interview = qs.filter(kind="INTERVIEW")

        return Response({
            "ok": True,
            "doc": ApplicationScoreSerializer(doc, many=True).data,
            "interview": ApplicationScoreSerializer(interview, many=True).data,
        }, status=200)

    def post(self, request, app_id: int):
        """
        ✅ Upsert:
        - 로그인한 채점자(request.user)가
        - 해당 application(app_id)에 대해
        - kind(DOC/INTERVIEW) 점수를 저장/갱신
        """
        app = get_object_or_404(Application, id=app_id)

        ser = ApplicationScoreUpsertSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"ok": False, "errors": ser.errors}, status=http_status.HTTP_400_BAD_REQUEST)

        data = ser.validated_data
        kind = data["kind"]

        score_obj, created = ApplicationScore.objects.update_or_create(
            application=app,
            reviewer=request.user,
            kind=kind,
            defaults={
                "score1": data.get("score1", 0),
                "score2": data.get("score2", 0),
                "score3": data.get("score3", 0),
                "comment": data.get("comment", ""),
            },
        )

        return Response({
            "ok": True,
            "created": created,
            "score": ApplicationScoreSerializer(score_obj).data,
        }, status=200)


class AdminApplicationDocFinalizeView(APIView):
    """
    ✅ 서류 결과 확정
    PATCH /api/applications/admin/<app_id>/doc-finalize
    body: { decision: "ACCEPTED" | "REJECTED" }
    """
    permission_classes = [IsInstructorOrStaff]

    @transaction.atomic
    def patch(self, request, app_id: int):
        app = get_object_or_404(Application.objects.select_related("user"), id=app_id)

        ser = AdminDecisionFinalizeSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"ok": False, "errors": ser.errors}, status=http_status.HTTP_400_BAD_REQUEST)

        decision = ser.validated_data["decision"]

        ok = app.finalize_doc(decision)
        if not ok:
            return Response({"ok": False, "errors": {"detail": "doc_decision already finalized or invalid decision"}}, status=400)

        app.save(update_fields=["doc_decision", "doc_finalized_at", "updated_at"])
        return Response({"ok": True, "doc_decision": app.doc_decision}, status=200)


class AdminApplicationFinalizeView(APIView):
    """
    ✅ 최종 결과 확정(면접 이후)
    PATCH /api/applications/admin/<app_id>/finalize
    body: { decision: "ACCEPTED" | "REJECTED" }

    규칙:
    - 서류 합격 확정(ACCEPTED)인 경우에만 최종 확정 가능
    - 최종 ACCEPTED 확정 시:
      - user.role = STUDENT
      - user.education_track = application.track
    """
    permission_classes = [IsInstructorOrStaff]

    @transaction.atomic
    def patch(self, request, app_id: int):
        app = get_object_or_404(Application.objects.select_related("user"), id=app_id)

        if app.doc_decision != "ACCEPTED":
            return Response({"ok": False, "errors": {"detail": "finalize requires doc_decision=ACCEPTED"}}, status=400)

        ser = AdminDecisionFinalizeSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"ok": False, "errors": ser.errors}, status=http_status.HTTP_400_BAD_REQUEST)

        decision = ser.validated_data["decision"]

        ok = app.finalize_final(decision)
        if not ok:
            return Response({"ok": False, "errors": {"detail": "final_decision already finalized or invalid decision"}}, status=400)

        # ✅ 최종 확정 반영
        u = app.user
        if decision == "ACCEPTED":
            u.role = "STUDENT"
            u.education_track = app.track
            u.save(update_fields=["role", "education_track"])
        else:
            # 불합격이면 role/track은 유지(원하면 APPLICANT로 강제도 가능)
            pass

        app.save(update_fields=["final_decision", "finalized_at", "updated_at"])
        return Response({"ok": True, "final_decision": app.final_decision}, status=200)


class AdminPersonalInterviewScheduleView(APIView):
    """
    ✅ 개별 면접 일정 설정
    PATCH /api/applications/admin/<app_id>/interview-schedule
    body: { personal_interview_datetime, personal_interview_location }
    """
    permission_classes = [IsInstructorOrStaff]

    def patch(self, request, app_id: int):
        app = get_object_or_404(Application, id=app_id)
        ser = PersonalInterviewScheduleSerializer(data=request.data)
        if not ser.is_valid():
            return Response({"ok": False, "errors": ser.errors}, status=400)
        app.personal_interview_datetime = ser.validated_data.get("personal_interview_datetime", "")
        app.personal_interview_location = ser.validated_data.get("personal_interview_location", "")
        app.save(update_fields=["personal_interview_datetime", "personal_interview_location", "updated_at"])
        return Response({"ok": True}, status=200)


@api_view(["GET"])
@permission_classes([AllowAny])
def track_application_settings(request):
    """공개용: 트랙별 지원 활성화 상태 조회 (인증 불필요)"""
    s = ResultNotificationSettings.get_settings()
    return Response({
        "ok": True,
        "tracks": {
            "PLANNING_DESIGN": s.track_planning_design_open,
            "FRONTEND": s.track_frontend_open,
            "BACKEND": s.track_backend_open,
            "AI_SERVER": s.track_ai_server_open,
        }
    })


class AdminResultNotificationSettingsView(APIView):
    """
    ✅ 합격 알림 설정 관리 (GET, PUT)
    GET /api/applications/admin/notification-settings
    PUT /api/applications/admin/notification-settings
    """
    permission_classes = [IsInstructorOrStaff]

    def get(self, request):
        settings = ResultNotificationSettings.get_settings()
        ser = ResultNotificationSettingsSerializer(settings)
        return Response({"ok": True, "settings": ser.data}, status=200)

    @transaction.atomic
    def put(self, request):
        settings = ResultNotificationSettings.get_settings()
        ser = ResultNotificationSettingsSerializer(settings, data=request.data, partial=True)
        if not ser.is_valid():
            return Response({"ok": False, "errors": ser.errors}, status=http_status.HTTP_400_BAD_REQUEST)
        
        # 수정자 기록
        settings.updated_by = request.user
        ser.save()

        return Response({"ok": True, "settings": ser.data}, status=200)


class AdminApplicationExportView(APIView):
    """
    ✅ 지원자 전체 목록 엑셀 다운로드
    GET /api/applications/admin/export
    - 기존 AdminApplicationListView와 동일한 필터/정렬 적용
    - 페이지네이션 없이 전체 반환
    """
    permission_classes = [IsInstructorOrStaff]

    def get(self, request):
        qs = (
            Application.objects
            .select_related("user")
            .prefetch_related("scores")
            .order_by("-updated_at")
        )

        st = request.query_params.get("status")
        tr = request.query_params.get("track")
        q = request.query_params.get("q")

        if st:
            qs = qs.filter(status=st)
        if tr:
            qs = qs.filter(track=tr)
        if q:
            qs = qs.filter(
                Q(user__email__icontains=q) |
                Q(user__name__icontains=q) |
                Q(user__student_id__icontains=q)
            )

        qs = qs.annotate(
            doc_avg=Coalesce(
                Avg("scores__score1", filter=Q(scores__kind="DOC"), output_field=FloatField()) +
                Avg("scores__score2", filter=Q(scores__kind="DOC"), output_field=FloatField()) +
                Avg("scores__score3", filter=Q(scores__kind="DOC"), output_field=FloatField()),
                None
            ),
            interview_avg=Coalesce(
                Avg("scores__score1", filter=Q(scores__kind="INTERVIEW"), output_field=FloatField()) +
                Avg("scores__score2", filter=Q(scores__kind="INTERVIEW"), output_field=FloatField()) +
                Avg("scores__score3", filter=Q(scores__kind="INTERVIEW"), output_field=FloatField()),
                None
            ),
            doc_count=Count("scores", filter=Q(scores__kind="DOC"), distinct=True),
            interview_count=Count("scores", filter=Q(scores__kind="INTERVIEW"), distinct=True),
        ).annotate(
            total_avg=Case(
                When(
                    doc_avg__isnull=False,
                    interview_avg__isnull=False,
                    then=(F("doc_avg") + F("interview_avg")) / Value(2.0),
                ),
                default=None,
                output_field=FloatField(),
            )
        )

        sort = request.query_params.get("sort")
        if sort == "TOTAL_DESC":
            qs = qs.order_by(F("total_avg").desc(nulls_last=True), "-updated_at")
        elif sort == "TOTAL_ASC":
            qs = qs.order_by(F("total_avg").asc(nulls_last=True), "-updated_at")

        TRACK_LABEL = {
            "PLANNING_DESIGN": "기획/디자인",
            "FRONTEND": "프론트엔드",
            "BACKEND": "백엔드",
            "AI_SERVER": "AI",
        }
        STATUS_LABEL = {
            "DRAFT": "작성중",
            "SUBMITTED": "제출완료",
            "ACCEPTED": "합격",
            "REJECTED": "불합격",
        }
        DECISION_LABEL = {
            "PENDING": "미확정",
            "ACCEPTED": "확정 합격",
            "REJECTED": "확정 불합격",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "지원자 목록"

        headers = [
            "순번", "이름", "학번", "학과", "연락처", "이메일",
            "지원 트랙", "상태", "제출일시",
            "서류 결과", "최종 결과",
            "서류 평균", "면접 평균", "총점 평균",
            "서류 채점수", "면접 채점수",
            "포트폴리오 URL",
            "자기소개 및 지원동기", "열정/성장 경험", "활동 계획 및 각오", "팀 협업 경험",
            "기획/디자인 경험", "사회 문제 해결 아이디어",
            "프로그래밍 학습 경험", "AI 서비스 인상",
            "웹 동작 원리", "코드 품질 기준",
            "UI/UX 개선 아이디어", "디자인 구현 우선순위",
        ]

        header_fill = PatternFill(fill_type="solid", fgColor="17538D")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        ws.row_dimensions[1].height = 20

        data_align = Alignment(vertical="top", wrap_text=True)

        def fmt(v):
            if v is None:
                return ""
            return round(float(v), 1)

        for row_idx, app in enumerate(qs, start=2):
            u = app.user
            submitted = app.submitted_at.strftime("%Y-%m-%d %H:%M") if app.submitted_at else ""

            row = [
                row_idx - 1,
                u.name,
                u.student_id,
                u.department,
                u.phone,
                u.email,
                TRACK_LABEL.get(app.track, app.track),
                STATUS_LABEL.get(app.status, app.status),
                submitted,
                DECISION_LABEL.get(getattr(app, "doc_decision", None) or "PENDING", "미확정"),
                DECISION_LABEL.get(getattr(app, "final_decision", None) or "PENDING", "미확정"),
                fmt(app.doc_avg),
                fmt(app.interview_avg),
                fmt(app.total_avg),
                app.doc_count or 0,
                app.interview_count or 0,
                app.portfolio_url or "",
                app.motivation or "",
                app.common_growth_experience or "",
                app.common_time_management or "",
                app.common_teamwork or "",
                app.planning_experience or "",
                app.planning_idea or "",
                app.ai_programming_level or "",
                app.ai_service_impression or "",
                app.backend_web_process or "",
                app.backend_code_quality or "",
                app.frontend_ui_experience or "",
                app.frontend_design_implementation or "",
            ]

            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = data_align

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="applicants.xlsx"'
        return response